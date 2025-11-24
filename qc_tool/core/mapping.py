# qc_tool/core/mapping.py
# Mapeo dinámico para QCPC con caché por mtime de plantilla (evita re-escaneo en cada corrida).
import importlib, unicodedata, json, os
from pathlib import Path
from openpyxl import load_workbook
from ..config.settings import ORACULOS_DIR

orig = importlib.import_module('guardar_excel')

def _norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if ord(c) < 128)  # quita tildes
    return s.strip().upper()

LABELS_PC = {
    "fecha_hora":       ["FECHA/HORA"],
    "realizado_por":    ["REALIZADO POR", "QC REALIZADO POR", "QC REALIZADO POR:"],
    "cliente":          ["CLIENTE"],
    "mother":           ["MOTHER"],
    "cpu":              ["CPU"],
    "gpu":              ["GPU", "GPU(S)"],
    "memoria_ram":      ["MEMORIA RAM"],
    "disco_duro_1":     ["DISCO DURO 1"],
    "disco_duro_2":     ["DISCO DURO 2"],
    "cd_dvd_rw":        ["CD / DVD RW", "LECTORA DVD"],
    "usb":              ["USB", "PUERTOS USB"],
    "cable_de_poder":   ["CABLE DE PODER", "CARGADOR", "CARGADOR/CABLE"],
    "hdmi":             ["HDMI"],
    "rj45":             ["RJ45"],
    "s_operativo":      ["S.OPERATIVO", "S.OPERATIVO /ACTIVACION", "S.OPERATIVO / ACTIVACION", "S.OPERATIVO / ACTIVACIÓN"],
    "drivers":          ["DRIVERS"],
    "office":           ["OFFICE"],
    "antivirus":        ["ANTIVIRUS"],
    "endpoint":         ["ENDPOINT CENTRAL"],
    "adobe_reader":     ["ADOBE READER"],
    "teamviewer":       ["TEAMVIEWER"],
    "7zip":             ["7ZIP", "7-ZIP"],
    "forti_vpn":        ["FORTI CLIENT VPN", "FORTICLIENT", "FORTICLIENT VPN"],
    "chrome":           ["CHROME", "GOOGLE CHROME"],
    "java":             ["JAVA"],
    "dominio":          ["DOMINIO", "UNIDO A DOMINIO"],
    "wifi":             ["WIFI", "WIFI FUNCIONANDO"],
}

SELLOS_PC = {
    "qc_rehecho":       ["QC REHECHO"],
    "sello_at_service": ["AT SERVICE", "AT-SERVICE", "AT SERVICE "],
    "micro_intel_amd":  ["MICRO INTEL/AMD", "MICRO INTEL / AMD"],
    "sello_garantia":   ["SELLO GARANTIA", "SELLO GARANTÍA"],
    "coa_windows":      ["COA WINDOWS"],
}

def _scan_column(ws, col_letter: str, terms: list[str], row_max: int = 120) -> int | None:
    wanted = set(_norm(t) for t in terms)
    for r in range(1, row_max + 1):
        val = ws[f"{col_letter}{r}"].value
        if _norm(val) in wanted:
            return r
    return None

def _build_map_pc_from_template(template_path: str) -> dict:
    try:
        wb = load_workbook(template_path, data_only=True)
        ws = wb.active
    except Exception:
        return getattr(orig, "MAP_PC", {})

    base = dict(getattr(orig, "MAP_PC", {}))
    result = {}

    for key, labels in LABELS_PC.items():
        r = _scan_column(ws, "B", labels)
        result[key] = base.get(key, ("C", 0)) if r is None else ("C", r)

    for key, labels in SELLOS_PC.items():
        r = _scan_column(ws, "G", labels)
        result[key] = base.get(key, ("G", 0)) if r is None else ("G", r)

    return result

def _load_cached_map(cache_file: Path) -> dict | None:
    try:
        with cache_file.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def _save_cached_map(cache_file: Path, data: dict) -> None:
    try:
        cache_file.parent.mkdir(parents=True, exist_ok=True)
        with cache_file.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _build_map_pc() -> dict:
    tpl = getattr(orig, "PLANTILLA_PC", None)
    if not tpl or not os.path.exists(tpl):
        return getattr(orig, "MAP_PC", {})

    mtime = os.path.getmtime(tpl)
    cache_file = Path(ORACULOS_DIR) / "_map_cache_pc.json"
    cached = _load_cached_map(cache_file)
    if cached and cached.get("template_mtime") == mtime:
        return cached.get("map_pc", {})

    mp = _build_map_pc_from_template(tpl)
    _save_cached_map(cache_file, {"template_mtime": mtime, "map_pc": mp})
    return mp

MAP_PC = _build_map_pc()
MAP_LAPTOP = getattr(orig, "MAP_LAPTOP", {})
for _name in ("FIELD_MAP", "NORMALIZED_MAP"):
    if hasattr(orig, _name):
        globals()[_name] = getattr(orig, _name)