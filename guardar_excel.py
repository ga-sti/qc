import unicodedata
from datetime import datetime
from openpyxl import load_workbook
import os
import socket

# --- Rutas a las plantillas ---
BASE_DIR         = os.path.dirname(__file__)
PLANTILLA_PC     = os.path.join(BASE_DIR, 'QCPC.xlsx')
PLANTILLA_LAPTOP = os.path.join(BASE_DIR, 'QCLAPTOP.xlsx')

# --- Etiquetas originales → claves internas ---
FIELD_MAP = {
    "FECHA/HORA":               "fecha_hora",
    "REALIZADO POR":            "realizado_por",
    "QC realizado por":         "realizado_por",
    "CLIENTE":                  "cliente",
    "AT SERVICE":               "sello_at_service",
    "MICRO INTEL/AMD":          "micro_intel_amd",
    "SELLO GARANTÍA":           "sello_garantia",
    "COA WINDOWS":              "coa_windows",
    "MOTHER":                   "mother",
    "CPU":                      "cpu",
    "GPU":                      "gpu",
    "GPU(s)":                   "gpu",
    "MEMORIA RAM":              "memoria_ram",
    "RAM Total (GB)":           "ram_total_gb",
    "RAM Slots usados":         "ram_slots_usados",
    "Tipo de RAM":              "tipo_de_ram",
    "DISCO DURO 1":             "disco_duro_1",
    "DISCO DURO 2":             "disco_duro_2",
    "Disco Total (GB)":         "disco_total_gb",
    "CD / DVD RW":              "cd_dvd_rw",
    "Lectora DVD":              "cd_dvd_rw",
    "USB":                      "usb",
    "Puertos USB":              "usb",
    "CARGADOR":                 "cable_de_poder",
    "Cargador/Cable":           "cable_de_poder",
    "CABLE DE PODER":           "cable_de_poder",
    "TECLADO":                  "teclado",
    "TECLADO (Testear)":        "teclado",
    "WEBCAM":                   "webcam",
    "HDMI":                     "hdmi",
    "RJ45":                     "rj45",
    "S.OPERATIVO":              "s_operativo",
    "S.OPERATIVO / ACTIVACIÓN": "s_operativo",
    "DRIVERS":                  "drivers",
    "Drivers OK":               "drivers",
    "OFFICE":                   "office",
    "ANTIVIRUS":                "antivirus",
    "ENDPOINT CENTRAL":         "endpoint",
    "ADOBE READER":             "adobe_reader",
    "TEAMVIEWER":               "teamviewer",
    "7ZIP":                     "7zip",
    "7-Zip":                    "7zip",
    "FORTI CLIENT VPN":         "forti_vpn",
    "FortiClient":              "forti_vpn",
    "CHROME":                   "chrome",
    "Google Chrome":            "chrome",
    "JAVA":                     "java",
    "DOMINIO":                  "dominio",
    "Unido a dominio":          "dominio",
    "WIFI":                     "wifi",
    "WiFi funcionando":         "wifi",
    "QC REHECHO":               "qc_rehecho",
}

# --- Coordenadas PC ---
MAP_PC = {
    "fecha_hora":       ("C",10),
    "realizado_por":    ("C",11),
    "cliente":          ("C",12),

    "mother":           ("C",15),
    "cpu":              ("C",16),
    "gpu":              ("C",17),
    "memoria_ram":      ("C",18),
    "disco_duro_1":     ("C",19),
    "disco_duro_2":     ("C",20),
    "cd_dvd_rw":        ("C",21),
    "usb":              ("C",22),
    "cable_de_poder":   ("C",23),
    "hdmi":             ("C",24),
    "rj45":             ("C",25),

    "s_operativo":      ("C",28),
    "drivers":          ("C",29),
    "office":           ("C",30),
    "antivirus":        ("C",31),
    "endpoint":         ("C",32),
    "adobe_reader":     ("C",33),
    "teamviewer":       ("C",34),
    "7zip":             ("C",35),
    "forti_vpn":        ("C",36),
    "chrome":           ("C",37),
    "java":             ("C",38),
    "dominio":          ("C",39),
    "wifi":             ("C",40),

    "qc_rehecho":       ("G",39),
    "sello_at_service": ("G",23),
    "micro_intel_amd":  ("G",24),
    "sello_garantia":   ("G",25),
    "coa_windows":      ("G",26),
}

# --- Coordenadas Laptop ---
MAP_LAPTOP = {
    "fecha_hora":       ("C",10),
    "realizado_por":    ("C",11),
    "cliente":          ("C",12),

    "mother":           ("C",15),
    "cpu":              ("C",16),
    "gpu":              ("C",17),
    "memoria_ram":      ("C",18),
    "disco_duro_1":     ("C",19),
    "disco_duro_2":     ("C",20),
    "cd_dvd_rw":        ("C",21),
    "usb":              ("C",22),
    "cable_de_poder":   ("C",23),
    "teclado":          ("C",24),
    "webcam":           ("C",25),
    "hdmi":             ("C",26),
    "rj45":             ("C",27),

    "s_operativo":      ("C",30),
    "drivers":          ("C",31),
    "office":           ("C",32),
    "antivirus":        ("C",33),
    "endpoint":         ("C",34),
    "adobe_reader":     ("C",35),
    "teamviewer":       ("C",36),
    "7zip":             ("C",37),
    "forti_vpn":        ("C",38),
    "chrome":           ("C",39),
    "java":             ("C",40),
    "dominio":          ("C",41),
    "wifi":             ("C",42),

    "qc_rehecho":       ("G",39),
    "sello_at_service": ("G",23),
    "micro_intel_amd":  ("G",24),
    "sello_garantia":   ("G",25),
    "coa_windows":      ("G",26),
}

def normalize_label(s: str) -> str:
    nk = unicodedata.normalize('NFKD', s)
    no_acc = "".join(c for c in nk if not unicodedata.combining(c))
    up = no_acc.upper()
    for ch in " ./-():":
        up = up.replace(ch, "")
    return up

NORMALIZED_MAP = {
    normalize_label(orig): key
    for orig, key in FIELD_MAP.items()
}

for _internal in set(FIELD_MAP.values()):
    NORMALIZED_MAP.setdefault(normalize_label(_internal), _internal)
    
def guardar_en_excel(datos_form, datos_sist, es_pc=True, output_path=None):
    plantilla = PLANTILLA_PC if es_pc else PLANTILLA_LAPTOP
    wb = load_workbook(plantilla)
    ws = wb.active
    coordmap = MAP_PC if es_pc else MAP_LAPTOP

    # 1) Traducir etiquetas de ambos diccionarios a claves internas
    raw = {**datos_form, **datos_sist}
    trad = {}
    for label, val in raw.items():
        key = NORMALIZED_MAP.get(normalize_label(label))
        if key:
            trad[key] = val
        elif normalize_label(label) not in ("TIPODEEQUIPO",):
            print(f"⚠️ Etiqueta NO mapeada: '{label}'")

    # 2) Fecha / hora siempre se genera al momento de crear el Excel
    trad["fecha_hora"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # 3) Armar memoria RAM unificada
    if "ram_total_gb" in trad:
        parts = []

        # Cantidad de RAM
        total = trad.pop("ram_total_gb")
        parts.append(f"{total} GB")

        # Slots usados (con singular/plural)
        if "ram_slots_usados" in trad:
            slots = trad.pop("ram_slots_usados")
            try:
                n = int(slots)
                suf = "slot" if n == 1 else "slots"
                parts.append(f"{n} {suf}")
            except Exception:
                # Si no es número, lo ponemos tal cual
                if slots:
                    parts.append(f"{slots} slots")

        # Tipo de RAM: solo si viene con un valor real (DDR4, etc.)
        if "tipo_de_ram" in trad:
            tipo = str(trad.pop("tipo_de_ram") or "").strip()
            if tipo and tipo.lower() not in ("desconocido", "unknown", "na", "n/a"):
                parts.append(tipo)

        trad["memoria_ram"] = " – ".join(parts)


    # 4) Disco total → DISCO DURO 1
    if "disco_total_gb" in trad:
        trad["disco_duro_1"] = f"{trad.pop('disco_total_gb')} GB"

    # 5) Volcar al Excel según el mapeo
    for key, (col, row) in coordmap.items():
        if key not in trad:
            continue
        val = trad[key]

        # --- MOTHER: texto + SI/NO ---
        if key == "mother":
            model = (val or "").strip()
            si_col = chr(ord(col) + 1)
            no_col = chr(ord(col) + 2)
            ws[f"{col}{row}"] = model if model else ""
            ws[f"{si_col}{row}"] = "X" if model else ""
            ws[f"{no_col}{row}"] = "" if model else "X"
            continue

        # --- DISCO DURO 2: lista o string + SI/NO ---
        if key == "disco_duro_2":
            txt = ", ".join(val) if isinstance(val, list) else (val or "")
            ws[f"{col}{row}"] = txt
            si_col = chr(ord(col) + 1)
            no_col = chr(ord(col) + 2)
            ws[f"{si_col}{row}"] = "X" if txt else ""
            ws[f"{no_col}{row}"] = "" if txt else "X"
            continue

        # --- CPU / GPU / MEMORIA / DISCO1 / USB: texto + SI ---
        if key in ("cpu", "gpu", "memoria_ram", "disco_duro_1", "usb"):
            txt = ", ".join(val) if isinstance(val, list) else val
            ws[f"{col}{row}"] = txt
            si_col = chr(ord(col) + 1)
            ws[f"{si_col}{row}"] = "X" if txt else ""
            continue

        # --- Campos de hardware con SI/NO ---
        if key in ("cd_dvd_rw", "cable_de_poder", "teclado", "webcam", "hdmi", "rj45"):
            si, no = chr(ord(col) + 1), chr(ord(col) + 2)
            if isinstance(val, str) and val.lower().startswith("s"):
                ws[f"{si}{row}"] = "X"
            else:
                ws[f"{no}{row}"] = "X"
            continue

        # --- Sellos (SI en H, NO en I) ---
        if key in ("sello_at_service", "micro_intel_amd", "sello_garantia", "coa_windows"):
            si, no = "H", "I"
            if isinstance(val, str) and val.lower().startswith("s"):
                ws[f"{si}{row}"] = "X"
            else:
                ws[f"{no}{row}"] = "X"
            continue

        # --- Dominio: texto + SI/NO ---
        if key == "dominio":
            si_col, no_col = chr(ord(col) + 1), chr(ord(col) + 2)
            dominio = (trad.get("dominio") or "").strip()
            if dominio:
                ws[f"{col}{row}"] = dominio
                ws[f"{si_col}{row}"] = "X"
                ws[f"{no_col}{row}"] = ""
            else:
                ws[f"{no_col}{row}"] = "X"
            continue

        # --- Software con SI/NO y posible texto ---
        if key in ("drivers", "wifi", "endpoint", "adobe_reader",
                   "teamviewer", "7zip", "forti_vpn", "chrome",
                   "java", "antivirus"):
            si_col, no_col = chr(ord(col) + 1), chr(ord(col) + 2)
            if isinstance(val, bool):
                ws[f"{si_col}{row}"] = "X" if val else ""
                ws[f"{no_col}{row}"] = "" if val else "X"
            elif isinstance(val, str):
                txt = val.strip()
                if txt.lower() in ("no", "no detectado", "error"):
                    ws[f"{no_col}{row}"] = "X"
                else:
                    ws[f"{col}{row}"] = txt
                    ws[f"{si_col}{row}"] = "X"
            elif isinstance(val, list):
                ws[f"{col}{row}"] = ", ".join(val)
                ws[f"{si_col}{row}"] = "X"
            continue

        # --- Office: versión + SI/NO ---
        if key == "office":
            si, no = chr(ord(col) + 1), chr(ord(col) + 2)
            if val:
                ws[f"{col}{row}"] = val
                ws[f"{si}{row}"] = "X"
            else:
                ws[f"{no}{row}"] = "X"
            continue

        # --- Sistema operativo: versión + SI/NO activación ---
        if key == "s_operativo":
            si_col, no_col = chr(ord(col) + 1), chr(ord(col) + 2)
            version, activated = ("", False)
            if isinstance(val, tuple) and len(val) == 2:
                version, activated = val
            elif isinstance(val, dict):
                version = val.get("version", "")
                activated = val.get("activated", False)
            elif isinstance(val, str):
                version = val
            ws[f"{col}{row}"] = version
            ws[f"{si_col}{row}"] = "X" if activated else ""
            ws[f"{no_col}{row}"] = "" if activated else "X"
            continue
        
        # 🚩 CUALQUIER OTRO (fecha_hora, realizado_por, cliente, etc.)
        ws[f"{col}{row}"] = val
    # Guardar archivo
    if not output_path:
        host = socket.gethostname()
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 1) Intentar usar la unidad donde está el script (pendrive)
        drive, _ = os.path.splitdrive(BASE_DIR)
        drive = drive.upper()

        if drive and drive != "C:":
            # Ej: E:\QC_Auto
            base_dest = os.path.join(drive + os.sep, "QC_Auto")
        # 2) Si no, intentar D:\ (muy típico de pendrive)
        elif os.path.exists(r"D:\\"):
            base_dest = os.path.join(r"D:\\", "QC_Auto")
        # 3) Último recurso: C:\QC_Auto
        else:
            base_dest = r"C:\QC_Auto"

        os.makedirs(base_dest, exist_ok=True)
        output_path = os.path.join(base_dest, f"{host}_{ts}.xlsx")
    else:
        # Si te pasan un output_path específico, respetalo
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb.save(output_path)
    return output_path

    # Asegurar que exista la carpeta destino
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb.save(output_path)
    return output_path
