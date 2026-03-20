import sys
import os
# Hack para que funcione el botón "Play" de VS Code sin problemas de rutas
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import unittest
from pathlib import Path
import openpyxl

from src.modelos import DatosFormulario, DatosSistema
from src.exportador import exportar_excel
from src.configuracion import DIR_PLANTILLAS

class TestEtapa3(unittest.TestCase):
    def setUp(self):
        """Prepara el entorno creando una plantilla falsa si no existe."""
        self.ruta_salida = Path("test_salida_qc.xlsx")
        self.plantilla_test = DIR_PLANTILLAS / "QCPC.xlsx"
        
        DIR_PLANTILLAS.mkdir(parents=True, exist_ok=True)
        
        # Evitamos que el test falle creando un Excel de prueba CON las etiquetas
        if not self.plantilla_test.exists():
            wb = openpyxl.Workbook()
            ws = wb.active
            # Inyectamos las etiquetas para que mapeo.py las encuentre
            ws["B12"] = "CLIENTE"
            ws["B16"] = "CPU"
            ws["B32"] = "OFFICE"
            wb.save(self.plantilla_test)

    def tearDown(self):
        """Limpia el archivo generado después del test."""
        if self.ruta_salida.exists():
            self.ruta_salida.unlink()

    def test_generacion_excel(self):
        print("\n[1/3] Preparando Mock de datos (Formulario + Sistema)...")
        form_mock = DatosFormulario(
            equipo="PC", 
            realizado_por="Gasti", 
            cliente="Empresa S.A."
        )
        sist_mock = DatosSistema(
            hostname="PC-TEST", 
            cpu="Intel Core i9-13900K", 
            ram_gb=32,
            office={"version": "Microsoft 365"},
            activado=True
        )

        print("[2/3] Ejecutando exportador...")
        ruta_generada = exportar_excel(form_mock, sist_mock, ruta_salida=self.ruta_salida)

        print(f"[3/3] Verificando Excel generado en: {ruta_generada}")
        self.assertTrue(Path(ruta_generada).exists(), "El archivo Excel no se creó.")

        wb = openpyxl.load_workbook(ruta_generada)
        ws = wb.active
        
        # Leemos todos los valores de la columna C (Textos) y D (Cruces de SI)
        valores_col_c = [ws[f"C{i}"].value for i in range(1, 50)]
        valores_col_d = [ws[f"D{i}"].value for i in range(1, 50)]
        
        # Verificamos que los datos se hayan escrito en ALGUNA parte de la columna correcta
        self.assertIn("Empresa S.A.", valores_col_c, "No se escribió el Cliente")
        self.assertIn("Intel Core i9-13900K", valores_col_c, "No se escribió el CPU")
        self.assertIn("Microsoft 365", valores_col_c, "No se escribió la versión de Office")
        
        # Verificamos que haya marcado al menos una "X" en la columna de SI (D)
        self.assertIn("X", valores_col_d, "No se marcó la cruz de SI en el software")
        
        print("\n¡Prueba de exportación a Excel completada con éxito y adaptación dinámica!")
        
if __name__ == "__main__":
    unittest.main()