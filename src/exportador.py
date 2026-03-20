# src/exportador.py
import os
import sys
import socket
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

# Importación de modelos y configuración
from src.modelos import DatosFormulario, DatosSistema
from src.configuracion import DIR_PLANTILLAS

# Mapeos actualizados según la imagen del Excel
MAP_PC = {
    "fecha_hora": ("C", 10), "realizado_por": ("C", 11), "cliente": ("C", 12),
    "mother": ("C", 15), "cpu": ("C", 16), "gpu": ("C", 17), "memoria_ram": ("C", 18),
    "disco_duro_1": ("C", 19), "disco_duro_2": ("C", 20),
    # USB y Sellos (Usan columnas I, J, K)
    "usb": ("I", 22), 
    "sello_at_service": ("I", 23), "micro_intel_amd": ("I", 24), 
    "sello_garantia": ("I", 25), "coa_windows": ("I", 26),
    # Pruebas HW (Columna C)
    "cable_de_poder": ("C", 23), "hdmi": ("C", 24), "rj45": ("C", 25),
    # Software
    "s_operativo": ("C", 28), "drivers": ("C", 29), "office": ("C", 30),
    "antivirus": ("C", 31), "endpoint": ("C", 32), "adobe_reader": ("C", 33),
    "teamviewer": ("C", 34), "7zip": ("C", 35), "forti_vpn": ("C", 36),
    "chrome": ("C", 37), "java": ("C", 38), "dominio": ("C", 39), "wifi": ("C", 40),
    "numero_serie": ("C", 41), "qc_rehecho": ("H", 38)
}

MAP_LAPTOP = {
    "fecha_hora": ("C", 10), "realizado_por": ("C", 11), "cliente": ("C", 12),
    "mother": ("C", 15), "cpu": ("C", 16), "gpu": ("C", 17), "memoria_ram": ("C", 18),
    "disco_duro_1": ("C", 19), "disco_duro_2": ("C", 20),
    # USB y Sellos
    "usb": ("I", 22), 
    "sello_at_service": ("I", 23), "micro_intel_amd": ("I", 24), 
    "sello_garantia": ("I", 25), "coa_windows": ("I", 26),
    # Pruebas HW
    "cable_de_poder": ("C", 23), "teclado": ("C", 24), "webcam": ("C", 25), "hdmi": ("C", 26), "rj45": ("C", 27),
    # Software
    "s_operativo": ("C", 30), "drivers": ("C", 31), "office": ("C", 32),
    "antivirus": ("C", 33), "endpoint": ("C", 34), "adobe_reader": ("C", 35),
    "teamviewer": ("C", 36), "7zip": ("C", 37), "forti_vpn": ("C", 38),
    "chrome": ("C", 39), "java": ("C", 40), "dominio": ("C", 41), "wifi": ("C", 42),
    "numero_serie": ("C", 43), "qc_rehecho": ("H", 38)
}

def obtener_unidades_destino() -> list[Path]:
    destinos = [Path("C:/")]
    try:
        unidad_actual = Path(sys.executable if getattr(sys, 'frozen', False) else __file__).anchor
        pendrive = Path(unidad_actual)
        if pendrive not in destinos: destinos.append(pendrive)
    except: pass
    return destinos

def generar_texto_log(datos_form: DatosFormulario, datos_sist: DatosSistema) -> str:
    ahora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    gpus = ", ".join(datos_sist.gpu) if datos_sist.gpu else "No detectada"
    avs = ", ".join(datos_sist.antivirus) if datos_sist.antivirus else "Windows Defender"
    hw_tests = "\n".join([f"   - {k.upper()}: {'[ OK ]' if v else '[ FALLA ]'}" for k, v in datos_form.hw.items()])
    
    return f"""
======================================================================
           REPORTE DE AUDITORÍA TÉCNICA - QC AUTOMATIZADO
======================================================================
CLIENTE: {datos_form.cliente} | TÉCNICO: {datos_form.realizado_por} | {ahora}
S.N: {datos_sist.numero_serie} | OS: {datos_sist.s_operativo}
----------------------------------------------------------------------
PRUEBAS:
{hw_tests}
USB: {datos_form.usb} puertos detectados.
ANTIVIRUS: {avs}
OFFICE: {datos_sist.office.get('version', 'No detectado')}
======================================================================
"""

def exportar_excel(datos_form: DatosFormulario, datos_sist: DatosSistema) -> str:
    # 1. Preparar datos
    ram_txt = f"{datos_sist.ram_gb} GB – {datos_sist.ram_slots} slots – {datos_sist.ram_tipo}"
    gpu_txt = ", ".join(datos_sist.gpu) if isinstance(datos_sist.gpu, list) else str(datos_sist.gpu)
    av_txt = ", ".join(datos_sist.antivirus) if datos_sist.antivirus else "Windows Defender"
    
    trad = {
        "fecha_hora": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "realizado_por": datos_form.realizado_por, "cliente": datos_form.cliente,
        "mother": datos_sist.mother, "cpu": datos_sist.cpu, "gpu": gpu_txt,
        "memoria_ram": ram_txt, "disco_duro_1": f"{datos_sist.disco_total_gb} GB",
        "s_operativo": (datos_sist.s_operativo, datos_sist.activado),
        "office": datos_sist.office.get("version") if datos_sist.office else "No detectado",
        "antivirus": av_txt, "dominio": datos_sist.dominio, "numero_serie": datos_sist.numero_serie,
        "usb": datos_form.usb
    }

    sw_raw = datos_sist.software or {}
    trad.update({
        "chrome": sw_raw.get("chrome"), "7zip": sw_raw.get("7zip"),
        "teamviewer": sw_raw.get("teamviewer"), "forti_vpn": sw_raw.get("forti_vpn"),
        "endpoint": sw_raw.get("endpoint"), "java": sw_raw.get("java"),
        "adobe_reader": sw_raw.get("adobe_reader"),
        "wifi": sw_raw.get("wifi") or (datos_form.sw.get("wifi") if datos_form.sw else False)
    })

    if datos_form.hw: trad.update(datos_form.hw)
    if datos_form.sellos: trad.update(datos_form.sellos)
    if datos_form.sw: trad.update({"drivers": datos_form.sw.get("drivers")})
    if datos_form.codigo_at: trad.update({"codigo_at": datos_form.codigo_at})

    # 2. Cargar Excel
    es_pc = (str(datos_form.equipo).upper() == "PC")
    nombre_tpl = "QCPC.xlsx" if es_pc else "QCLAPTOP.xlsx"
    wb = load_workbook(DIR_PLANTILLAS / nombre_tpl)
    ws = wb.active
    coordmap = MAP_PC if es_pc else MAP_LAPTOP

    # 3. Llenado de celdas inteligente
    for key, (col, row) in coordmap.items():
        if key not in trad: continue
        val = trad[key]
        
        # --- COLUMNA C (Texto/Valores generales) ---
        if col == "C":
            if key == "s_operativo":
                ver, act = val if isinstance(val, tuple) else (val, False)
                ws[f"C{row}"] = str(ver)
                ws[f"D{row}"], ws[f"E{row}"] = ("X", "") if act else ("", "X")
            elif key == "numero_serie":
                ws[f"C{row}"] = str(val or "No detectado")
                if val: ws[f"D{row}"] = "X"
            elif isinstance(val, bool) or key in ("endpoint", "chrome", "7zip", "teamviewer", "adobe_reader", "forti_vpn", "java"):
                # Apps: Marcamos X en SI (D) o NO (E)
                if val: ws[f"D{row}"] = "X"
                else: ws[f"E{row}"] = "X"
            elif key == "antivirus":
                ws[f"C{row}"] = str(val)
                ws[f"D{row}"] = "X" # Si el motor lo detectó, marcamos que SI
            else:
                ws[f"C{row}"] = str(val or "")
                # Marcamos X en D para Hardware que dio OK
                if val and key not in ("disco_duro_2", "fecha_hora", "realizado_por", "cliente", "mother", "cpu", "gpu", "memoria_ram", "disco_duro_1"):
                    ws[f"D{row}"] = "X"

        # --- COLUMNA I (SI / NO / CARACTERÍSTICA) ---
        elif col == "I":
            if key == "usb":
                # 🟢 Lógica especial USB: Cantidad en K, X en I
                ws[f"K{row}"] = f"{val} PUERTOS"
                if val > 0: ws[f"I{row}"] = "X"
                else: ws[f"J{row}"] = "X"
            else:
                # 🟢 Sellos (Filas 23-26)
                if val: 
                    ws[f"I{row}"] = "X"
                    # Si es el sello de AT Service, ponemos el código en K
                    if key == "sello_at_service" and trad.get("codigo_at"):
                        ws[f"K{row}"] = f"COD: {trad['codigo_at']}"
                else: 
                    ws[f"J{row}"] = "X"

        # --- COLUMNA H (QC Rehecho) ---
        elif col == "H":
            is_ok = val if isinstance(val, bool) else str(val).lower().startswith("s")
            ws[f"H{row}"], ws[f"I{row}"] = ("X", "") if is_ok else ("", "X")

    # 4. Guardado
    fecha_slug = datetime.now().strftime("%Y%m%d")
    cliente_slug = "".join(c for c in str(datos_form.cliente) if c.isalnum()).upper()
    nombre_archivo = f"QC_{fecha_slug}_{cliente_slug}_{datos_sist.hostname}"
    
    unidades = obtener_unidades_destino()
    ruta_retorno = ""
    for unidad in unidades:
        base_path = unidad / "QC" / "Resultado"
        (base_path / "Excels").mkdir(parents=True, exist_ok=True)
        (base_path / "Logs").mkdir(parents=True, exist_ok=True)

        final_excel = base_path / "Excels" / f"{nombre_archivo}.xlsx"
        wb.save(final_excel)
        
        with (base_path / "Logs" / f"{nombre_archivo}.txt").open("w", encoding="utf-8") as f:
            f.write(generar_texto_log(datos_form, datos_sist))

        if str(unidad).startswith("C"): ruta_retorno = str(final_excel)

    return ruta_retorno